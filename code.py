from openpyxl import Workbook
import time
import requests
from bs4 import BeautifulSoup, SoupStrainer


# ---------------------------------------------------------------------------------------------------------------------
# Working with EXCEL
# ---------------------------------------------------------------------------------------------------------------------

workbook = Workbook()
worksheet = workbook.active
rent_worksheet = workbook.create_sheet('Rent', 0)
sale_worksheet = workbook.create_sheet('Sale', 1)

table_header = ["Price £", "Price £ per sq ft", "Price £ per calendar month", "Address", "City", "Region", "Postal code",
                "URL of the add", "Search identifier", "Description", "Surface sq ft", "Date of scraping",
                "Company that propose"]
rent_worksheet.append(table_header)
sale_worksheet.append(table_header)

# ---------------------------------------------------------------------------------------------------------------------
# Working with requests
# ---------------------------------------------------------------------------------------------------------------------

session = requests.Session()
session.headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                                 '(KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36'}


def load_page(url, parameters):
    while True:
        try:
            request = session.get(url=url, params=parameters)
            session.close()
            return request
        except:
            time.sleep(5)


def load_page_text(url, parameters):
    while True:
        try:
            request = session.get(url=url, params=parameters)
            session.close()
            return request.text
        except:
            time.sleep(5)


def get_detail_url(list_of_url_endings):
    main_url = 'https://www.zoopla.co.uk/'
    list_for_full_urls = []
    for each_ending in list_of_url_endings:
        url = main_url + each_ending
        list_for_full_urls.append(url)
    return list_for_full_urls


# ---------------------------------------------------------------------------------------------------------------------
# Working with scraper
# ---------------------------------------------------------------------------------------------------------------------

def soup_page(requested_page, parsing_limits=None):
    xml_text = BeautifulSoup(requested_page, 'lxml', parse_only=SoupStrainer(name=parsing_limits[0], attrs=parsing_limits[1]))
    return xml_text


def find_tag(xml_text, search_options):
    tag = xml_text.find(name=search_options[0], attrs=search_options[1])
    return tag


def find_tags(xml_text, search_options):
    tags = xml_text.find_all(name=search_options[0], attrs=search_options[1])
    return tags


def get_value(tag, search_value_in):
    if tag is None:
        value = None
    elif tag.name == search_value_in:
        value = tag.text
    else:
        value = tag.get(search_value_in)
    return value


def page_parsing(url, url_parameters, parsing_limits, search_options, value_options):
    requesting_text = load_page_text(url, url_parameters)
    souping = soup_page(requesting_text, parsing_limits)
    found_tags = find_tags(souping, search_options)
    parsing_result = []
    for tag in found_tags:
        value = get_value(tag, value_options)
        parsing_result.append(value)
    return parsing_result


def detail_parsing(url, url_parameters, parsing_limits, search_options, value_options):
    requesting_page = load_page(url, url_parameters)
    souping = soup_page(requesting_page.text, parsing_limits)
    requesting_date = requesting_page.headers['Date']
    extended_info = [url, requesting_date]
    values_result = []
    for each_option in search_options:
        found_tag = find_tag(souping, each_option)
        found_value = get_value(found_tag, value_options[search_options.index(each_option)])
        values_result.append(found_value)
    detail_parsing_result = values_result + extended_info
    return detail_parsing_result


# ---------------------------------------------------------------------------------------------------------------------
# Working with data
# ---------------------------------------------------------------------------------------------------------------------

def price_corrector(raw_price, raw_price_pft, raw_price_pcm):
    if raw_price.isalpha() or raw_price is None:
        corrected_price = raw_price
    else:
        corrected_price = raw_price[1:-3]
    if raw_price_pft is None:
        corrected_price_pft = raw_price_pft
    else:
        corrected_price_pft = raw_price_pft[2:-12]
    if raw_price_pcm is None:
        corrected_price_pcm = raw_price_pcm
    else:
        corrected_price_pcm = raw_price_pcm[1: -4]
    return corrected_price, corrected_price_pft, corrected_price_pcm


def address_converter(address):
    con_address, con_city, con_postal_code = None, None, None
    if address is not None:
        street = str()
        r_address = address.replace("Greater ", "")
        address_pieces = r_address.split(" ")
        if 0 < len(address_pieces) < 2:
            con_postal_code = address_pieces[-1]
        elif 1 < len(address_pieces) < 3:
            con_postal_code = address_pieces[-1]
            con_city = address_pieces[-2]
        elif len(address_pieces) > 2:
            con_postal_code = address_pieces[-1]
            con_city = address_pieces[-2]
            for each_piece in address_pieces:
                if address_pieces.index(each_piece) == address_pieces.index(address_pieces[-2]):
                    break
                street += " " + each_piece
                con_address = street[1:-1]
    return con_address, con_city, con_postal_code


def get_identifier(url):
    first_split = url.split("?")
    second_split = first_split[0].split("/")
    identifier = second_split[-1]
    return identifier


def description_corrector(desc):
    if desc is None:
        return None
    replacing_desc = desc.replace("\n", "")
    striping_desc = replacing_desc.strip()
    return striping_desc


def surface_corrector(surf):
    if surf is None:
        return None
    rep_surf = surf.replace("From", "")
    r_surf = rep_surf.replace(" sq. ft", "")
    return r_surf


def value_sort(detail_values):
    price, price_pft, price_pcm = price_corrector(detail_values[0], detail_values[1], detail_values[2])
    address, city, postal_code = address_converter(detail_values[3])
    description = description_corrector(detail_values[5])
    surface = surface_corrector(detail_values[4])
    agent, url, date_of_scraping, region = detail_values[6], detail_values[7], detail_values[8], detail_values[9]
    search_identifier = get_identifier(url)
    sorted_list = [price, price_pft, price_pcm, address, city, region, postal_code, url, search_identifier,
                   description, surface, date_of_scraping, agent]
    return sorted_list


# ---------------------------------------------------------------------------------------------------------------------
# Running program
# ---------------------------------------------------------------------------------------------------------------------

PAGE_AMOUNT = [None,
               ['div', {'class': 'paginate bg-muted'}],
               ['a', {'href': True}],
               'a']

PROPOSE_URLS = [{'page_size': 25, 'pn': 1},
                ['ul', {'class': 'listing-results clearfix js-gtm-list'}],
                ['a', 'listing-results-price text-price'],
                'href']

PROPOSE_DETAILS = [None,
                   ['div', {'class': 'ui-layout'}],
                   [['p', 'ui-pricing__main-price ui-text-t4'],
                    ['p', 'ui-pricing__area-price'],
                    ['p', 'ui-pricing__alt-price'],
                    ['h2', 'ui-property-summary__address'],
                    ['span', 'dp-features-list__text'],
                    ['div', 'dp-description__text'],
                    ['h4', 'ui-agent__name']],
                   ['p', 'p', 'p', 'h2', 'span', 'div', 'h4']
                   ]


REGIONS = [['Greater London', 'South East England', 'East Midlands', 'East of England', 'North East England',
            'North West England', 'South West England', 'West Midlands', 'Yorkshire and The Humber', 'Isle of Man',
            'Channel Isles', 'Scotland', 'Wales', 'Northern Ireland'],
           ['london', 'south-east-england', 'east-midlands', 'east-of-england', 'north-east-england',
            'north-west-england', 'south-west-england', 'west-midlands', 'yorkshire-and-the-humber', 'isle-of-man',
            'channel-isles', 'scotland', 'wales', 'northern-ireland']]


def run_program(page_amount, propose_urls, propose_details, regions):
    page_url_parameters = propose_urls[0]
    sheets_name = ['Rent', 'Sale']
    propose_types = ['to-rent', 'for-sale']
    for propose_type in propose_types:
        sheet_name = sheets_name[propose_types.index(propose_type)]
        excel_sheet = workbook[sheet_name]
        propose_type_url = ('https://www.zoopla.co.uk/%s/commercial/offices/' % propose_type)
        list_of_regions = regions[0]
        region_urls = regions[1]
        for region_url in region_urls:
            region = list_of_regions[region_urls.index(region_url)]
            url = propose_type_url + region_url
            amount_of_pages = page_parsing(url=url,
                                           url_parameters=page_amount[0],
                                           parsing_limits=page_amount[1],
                                           search_options=page_amount[2],
                                           value_options=page_amount[3])
            if len(amount_of_pages) == 0:
                last_page = 1
            else:
                last_page = int(amount_of_pages[-2])
            # ___________________________________________
            start_page = 1
            while start_page < last_page:
                page_url_parameters['pn'] = start_page
                detail_url_ends = page_parsing(url=url,
                                               url_parameters=page_url_parameters,
                                               parsing_limits=propose_urls[1],
                                               search_options=propose_urls[2],
                                               value_options=propose_urls[3])
                detail_urls = get_detail_url(detail_url_ends)
                for each_url in detail_urls:
                    propose_info = detail_parsing(url=each_url,
                                                  url_parameters=propose_details[0],
                                                  parsing_limits=propose_details[1],
                                                  search_options=propose_details[2],
                                                  value_options=propose_details[3])
                    propose_info.append(region)
                    print(propose_info)
                    sorted_values = value_sort(propose_info)
                    sorted_values.append(page_url_parameters['pn'])
                    print(sorted_values)
                    excel_sheet.append(sorted_values)
                    workbook.save("scraper zoopla1.xlsx")
                start_page += 1


if __name__ == '__main__':
    run_program(PAGE_AMOUNT, PROPOSE_URLS, PROPOSE_DETAILS, REGIONS)
